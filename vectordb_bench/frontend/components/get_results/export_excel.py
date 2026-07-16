"""
Export benchmark results to a formatted Excel file.
Captures the same information as the results page: summary, configurations per instance, and metrics per case.
Includes a Glossary sheet and cell comments on column headers so readers can understand each metric and parameter.
"""

import io
import warnings
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from vectordb_bench.frontend.components.check_results.results_summary import (
    get_databases_and_tuning,
    get_query_run_counts,
)
from vectordb_bench.frontend.config.parameter_tooltips import PARAM_TOOLTIPS, get_case_param_tooltip
from vectordb_bench.frontend.config.results_metric_tooltips import (
    get_results_metric_display_name,
    get_results_metric_group_order,
    get_results_metric_tooltip,
    group_metrics_for_display,
)
from vectordb_bench.frontend.components.get_results.theoretical_estimates import (
    compute_theoretical_breakdown,
)
from vectordb_bench.metric import metric_unit_map

# Author label for Excel cell comments (shown in comment popup)
_COMMENT_AUTHOR = "VectorDBBench"

# Metrics whose stored unit is a data size. Values are auto-scaled to the most
# readable unit per cell (e.g. 2048 MB -> "2.00 GB"). The key is the metric's
# base unit as declared in metric_unit_map; the value is that unit's size in bytes.
_SIZE_UNIT_BASE_BYTES = {
    "bytes": 1,
    "kb": 1024,
    "mb": 1024**2,
    "gb": 1024**3,
}


def _size_metric_base_bytes(metric: str) -> int | None:
    """Return the byte-size of one unit for a size metric, or None if not a size metric."""
    return _SIZE_UNIT_BASE_BYTES.get(str(metric_unit_map.get(metric, "")).strip().lower())


def _human_readable_size(num_bytes: float) -> str:
    """Format a byte count into the largest unit whose value is < 1024 (1024-based steps)."""
    units = ["bytes", "KB", "MB", "GB", "TB", "PB"]
    size = float(num_bytes)
    idx = 0
    while abs(size) >= 1024 and idx < len(units) - 1:
        size /= 1024
        idx += 1
    if idx == 0:
        return f"{int(round(size))} {units[idx]}"
    return f"{size:.2f} {units[idx]}"

# Excel sheet names must be <= 31 characters
EXCEL_SHEET_TITLE_MAX_LEN = 31


def _sheet_title(name: str, fallback: str = "Sheet") -> str:
    """Return a safe Excel sheet title (max 31 chars, no invalid chars)."""
    safe = "".join(c if c.isalnum() or c in " &-_" else "_" for c in (name or "")).strip() or fallback
    if len(safe) > EXCEL_SHEET_TITLE_MAX_LEN:
        safe = safe[:EXCEL_SHEET_TITLE_MAX_LEN]
    return safe


def _readable_param_key(key: str) -> str:
    return key.replace("_", " ").title()


def _metric_value(d: dict, metric: str):
    """Return a value suitable for Excel; flatten list metrics to first element or comma-sep."""
    v = d.get(metric, None)
    if v is None:
        return ""
    if isinstance(v, list):
        if len(v) == 0:
            return ""
        first = v[0]
        while isinstance(first, list) and first:
            first = first[0]
        if isinstance(first, (int, float)):
            return float(first)
        return ", ".join(str(x) for x in v[:10]) + ("..." if len(v) > 10 else "")
    if isinstance(v, (int, float)):
        return float(v) if isinstance(v, float) else v
    return v


def _style_header(cell):
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )


def _style_cell(cell):
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    cell.alignment = Alignment(vertical="center")


def _set_cell_comment(cell, text: str, max_len: int = 500) -> None:
    """Attach a comment to a cell so readers see the description on hover. Truncate if very long."""
    if not text or not text.strip():
        return
    clean = text.strip().replace("\n", " ")
    if len(clean) > max_len:
        clean = clean[: max_len - 3] + "..."
    cell.comment = Comment(clean, _COMMENT_AUTHOR)


def _write_glossary_sheet(wb) -> None:
    """Add a Glossary sheet explaining every result metric and configuration parameter in plain language."""
    # Insert at index 1 so Glossary appears right after Summary
    ws = wb.create_sheet(_sheet_title("Glossary", "Glossary"), 1)
    row = 1
    # Intro
    ws.cell(row=row, column=1, value=(
        "This sheet explains every column heading used in this workbook. "
        "You can also hover over any column header in the Summary or case sheets to see the same description in a pop-up note."
    ))
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 2

    # --- How many searches run (plain language) ---
    ws.cell(row=row, column=1, value="How search tests work")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="Term")
    ws.cell(row=row, column=2, value="Plain explanation")
    _style_header(ws.cell(row=row, column=1))
    _style_header(ws.cell(row=row, column=2))
    row += 1
    _search_workload_glossary = [
        (
            "Serial search",
            "Searches run one after another, in order. The number of searches equals the number of "
            "query vectors in the test set (from the dataset test file). Example: 1,000 vectors → "
            "1,000 searches for that pass. This pass is used for recall, NDCG, and serial latency (p99 / p95).",
        ),
        (
            "Concurrent search",
            "Many searches run in parallel for a fixed amount of time (default 30 seconds per concurrency "
            "level in a standard performance case). The benchmark keeps cycling through the same query vectors "
            "until the timer stops. So the total count is not “one search per vector”; it can be much higher "
            "or lower depending on speed. Each concurrency level (e.g. 1 worker, then 5, then 10) is its own timed run.",
        ),
        (
            "Streaming / search while loading",
            "Same pattern: one full serial pass over every test vector, plus timed concurrent search. "
            "The concurrent phase often uses a longer time per level than standard performance (set in the case).",
        ),
    ]
    for term, desc in _search_workload_glossary:
        ws.cell(row=row, column=1, value=term)
        ws.cell(row=row, column=2, value=desc)
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
        _style_cell(ws.cell(row=row, column=1))
        _style_cell(ws.cell(row=row, column=2))
        row += 1
    row += 2

    # --- Result metrics (from RESULTS_METRIC_GROUPS) ---
    ws.cell(row=row, column=1, value="Result metrics (column headings in case sheets)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="Metric name")
    ws.cell(row=row, column=2, value="What it means (in plain language)")
    _style_header(ws.cell(row=row, column=1))
    _style_header(ws.cell(row=row, column=2))
    row += 1
    seen_metrics = set()
    for _group_heading, metric_list in get_results_metric_group_order():
        for metric in metric_list:
            if metric in seen_metrics:
                continue
            seen_metrics.add(metric)
            name = get_results_metric_display_name(metric)
            desc = get_results_metric_tooltip(metric)
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=desc)
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
            _style_cell(ws.cell(row=row, column=1))
            _style_cell(ws.cell(row=row, column=2))
            row += 1
    row += 2

    # --- Configuration parameters ---
    ws.cell(row=row, column=1, value="Configuration parameters (Summary and per-instance tuning)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="Parameter name")
    ws.cell(row=row, column=2, value="What it means (in plain language)")
    _style_header(ws.cell(row=row, column=1))
    _style_header(ws.cell(row=row, column=2))
    row += 1
    seen_readable = set()
    for param_key in sorted(PARAM_TOOLTIPS.keys()):
        readable = _readable_param_key(param_key)
        if readable in seen_readable:
            continue
        seen_readable.add(readable)
        desc = get_case_param_tooltip(param_key)
        if not desc:
            continue
        ws.cell(row=row, column=1, value=readable)
        ws.cell(row=row, column=2, value=desc)
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
        _style_cell(ws.cell(row=row, column=1))
        _style_cell(ws.cell(row=row, column=2))
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 90


def _write_component_usage_sheet(wb, shown_data: list[dict]) -> None:
    """One row per (instance, case, component): disk / RAM / cached / expected-cache usage.

    Data comes from Metric.db_component_usage_json, reported by the database itself
    (e.g. Qdrant's collection memory API). Sheet is only added when at least one
    result carries a breakdown.
    """
    import json as _json

    rows = []
    for d in shown_data:
        raw = d.get("db_component_usage_json") or ""
        if not raw:
            continue
        try:
            breakdown = _json.loads(raw)
        except (ValueError, TypeError):
            continue
        instance = d.get("bar_display_name") or d.get("db_name", "")
        case_name = d.get("case_name", "")
        source = breakdown.get("source", "")
        for comp in breakdown.get("components", []):
            rows.append((instance, case_name, source, comp))
    if not rows:
        return

    ws = wb.create_sheet(_sheet_title("Component Usage", "Component Usage"))
    ws.cell(row=1, column=1, value=(
        "Per-component storage and memory usage reported by the database itself after the run "
        "(same data as the Memory tab in Qdrant's web UI). "
        "RAM = non-evictable heap memory; Cached = file pages resident in the OS page cache (evictable); "
        "Expected cache = data that should ideally be cached for best performance; "
        "Disk = total file sizes on disk."
    ))
    ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.row_dimensions[1].height = 45

    headers = ["Instance", "Case", "Component", "Disk", "RAM", "Cached", "Expected cache", "Source"]
    header_row = 3
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        _style_header(cell)

    def _fmt(comp: dict, key: str):
        v = comp.get(key)
        # Missing key = not measured (e.g. telemetry fallback); leave blank rather than 0.
        return _human_readable_size(v) if isinstance(v, (int, float)) else ""

    r = header_row + 1
    for instance, case_name, source, comp in rows:
        component_name = comp.get("component", "")
        size_est = comp.get("size_bytes")
        values = [
            instance,
            case_name,
            component_name if size_est is None else f"{component_name} (size: {_human_readable_size(size_est)})",
            _fmt(comp, "disk_bytes"),
            _fmt(comp, "ram_bytes"),
            _fmt(comp, "cached_bytes"),
            _fmt(comp, "expected_cache_bytes"),
            source,
        ]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            _style_cell(cell)
        r += 1

    widths = [22, 26, 40, 14, 14, 14, 16, 46]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w


def _write_theoretical_sheet(wb, shown_data: list[dict]) -> None:
    """Best/expected/worst-case RAM and disk estimates per component, per instance x case.

    Computed from dataset size/dimension and each instance's tuning parameters so readers
    can compare theory against the measured metrics and the Component Usage sheet.
    """
    # One block per unique (instance, case, dataset, tuning); skip rows lacking dataset info.
    blocks = []
    seen = set()
    for d in shown_data:
        n = int(d.get("dataset_size") or 0)
        dim = int(d.get("dataset_dim") or 0)
        if not n or not dim:
            continue
        instance = d.get("bar_display_name") or d.get("db_name", "")
        case_name = d.get("case_name", "")
        key = (instance, case_name)
        if key in seen:
            continue
        seen.add(key)
        tuning = {k: getattr(v, "value", v) for k, v in (d.get("tuning_params") or {}).items()}
        breakdown = compute_theoretical_breakdown(n, dim, tuning)
        if breakdown:
            blocks.append((instance, case_name, d.get("dataset_name", ""), breakdown))
    if not blocks:
        return

    ws = wb.create_sheet(_sheet_title("Theoretical Estimates", "Theoretical Estimates"))
    ws.cell(row=1, column=1, value=(
        "Theoretical resource estimates computed from dataset size, dimension, and each instance's "
        "configuration. Compare against measured avg/peak memory, DB data dir bytes written, and the "
        "Component Usage sheet. Best = minimal steady-state; Expected = typical incl. overhead; "
        "Worst = peak during indexing/optimization (use for capacity planning)."
    ))
    ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws.row_dimensions[1].height = 45

    headers = [
        "Component",
        "RAM (best)", "RAM (expected)", "RAM (worst)",
        "Disk (best)", "Disk (expected)", "Disk (worst)",
        "How it's calculated",
    ]
    r = 3
    for instance, case_name, dataset_name, breakdown in blocks:
        title = f"{instance} — {case_name}" + (f" ({dataset_name})" if dataset_name else "")
        ws.cell(row=r, column=1, value=title).font = Font(bold=True, size=12)
        r += 1
        for c, h in enumerate(headers, start=1):
            _style_header(ws.cell(row=r, column=c, value=h))
        r += 1
        for comp in breakdown["components"]:
            values = [
                comp["component"],
                *[_human_readable_size(v) if v else "—" for v in comp["ram"]],
                *[_human_readable_size(v) if v else "—" for v in comp["disk"]],
                comp["note"],
            ]
            for c, v in enumerate(values, start=1):
                _style_cell(ws.cell(row=r, column=c, value=v))
            ws.cell(row=r, column=8).alignment = Alignment(wrap_text=True, vertical="center")
            r += 1
        totals = breakdown["totals"]
        total_values = [
            "TOTAL",
            *[_human_readable_size(v) for v in totals["ram"]],
            *[_human_readable_size(v) for v in totals["disk"]],
            "",
        ]
        for c, v in enumerate(total_values, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            _style_cell(cell)
            cell.font = Font(bold=True)
        r += 1
        for line in breakdown["assumptions"]:
            cell = ws.cell(row=r, column=1, value=f"Note: {line}")
            cell.alignment = Alignment(wrap_text=True)
            cell.font = Font(italic=True, size=9)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            r += 1
        r += 2  # blank rows between blocks

    widths = [28, 14, 15, 14, 14, 15, 14, 80]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w


def build_results_excel(shown_data: list[dict], failed_tasks: dict) -> bytes:
    """
    Build a formatted Excel workbook from the same data shown on the results page.
    Supports unified results with multiple benchmark types (Standard, Int Filter, Label Filter).
    Returns the file as bytes for use with st.download_button.
    """
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message=".*31 characters.*")
        wb = Workbook()
        wb.remove(wb.active)

        # --- Summary sheet ---
        ws_summary = wb.create_sheet(_sheet_title("Summary", "Summary"), 0)
        row = 1
        db_tuning = get_databases_and_tuning(shown_data)
        counts = get_query_run_counts(shown_data)
        benchmark_types = sorted(set(d.get("benchmark_type", "Standard Performance") for d in shown_data))

        if len(benchmark_types) > 1:
            ws_summary.cell(row=row, column=1, value="Benchmark types in this export")
            ws_summary.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            ws_summary.cell(row=row, column=1, value=", ".join(benchmark_types))
            row += 2
        ws_summary.cell(row=row, column=1, value="Databases in this run")
        ws_summary.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        if db_tuning:
            ws_summary.cell(row=row, column=1, value=", ".join(db_tuning.keys()))
            row += 2
            ws_summary.cell(row=row, column=1, value="Configurations per instance")
            ws_summary.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            param_keys = sorted(set(k for params in db_tuning.values() for k in params.keys()))
            if param_keys:
                ws_summary.cell(row=row, column=1, value="Instance")
                _style_header(ws_summary.cell(row=row, column=1))
                _set_cell_comment(
                    ws_summary.cell(row=row, column=1),
                    "Name of the database instance in this run (e.g. Clickhouse 1, Milvus 2). Each instance can have its own configuration.",
                )
                for c, k in enumerate(param_keys, start=2):
                    ws_summary.cell(row=row, column=c, value=_readable_param_key(k))
                    _style_header(ws_summary.cell(row=row, column=c))
                    desc = get_case_param_tooltip(k)
                    if desc:
                        _set_cell_comment(ws_summary.cell(row=row, column=c), desc)
                row += 1
                for instance_name, params in db_tuning.items():
                    ws_summary.cell(row=row, column=1, value=instance_name)
                    _style_cell(ws_summary.cell(row=row, column=1))
                    for c, k in enumerate(param_keys, start=2):
                        v = params.get(k, "")
                        v = getattr(v, "value", v) if v != "" else "—"
                        ws_summary.cell(row=row, column=c, value=str(v))
                        _style_cell(ws_summary.cell(row=row, column=c))
                    row += 1
            else:
                for instance_name in db_tuning:
                    ws_summary.cell(row=row, column=1, value=instance_name)
                    ws_summary.cell(row=row, column=2, value="—")
                    _style_cell(ws_summary.cell(row=row, column=1))
                    _style_cell(ws_summary.cell(row=row, column=2))
                    row += 1
            row += 1
        ws_summary.cell(row=row, column=1, value="Queries run")
        ws_summary.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        if counts:
            ws_summary.cell(row=row, column=1, value="Case")
            ws_summary.cell(row=row, column=2, value="Count")
            _style_header(ws_summary.cell(row=row, column=1))
            _style_header(ws_summary.cell(row=row, column=2))
            row += 1
            for case_name in sorted(counts.keys()):
                n = counts[case_name]
                ws_summary.cell(row=row, column=1, value=case_name)
                ws_summary.cell(row=row, column=2, value=n)
                _style_cell(ws_summary.cell(row=row, column=1))
                _style_cell(ws_summary.cell(row=row, column=2))
                row += 1
        if failed_tasks:
            row += 1
            ws_summary.cell(row=row, column=1, value="Failed / Timeout (by case)")
            ws_summary.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            for case_name, db_labels in failed_tasks.items():
                for db_name, label in db_labels.items():
                    ws_summary.cell(row=row, column=1, value=case_name)
                    ws_summary.cell(row=row, column=2, value=db_name)
                    ws_summary.cell(row=row, column=3, value=str(label))
                    _style_cell(ws_summary.cell(row=row, column=1))
                    _style_cell(ws_summary.cell(row=row, column=2))
                    _style_cell(ws_summary.cell(row=row, column=3))
                    row += 1
        for col in range(1, 20):
            ws_summary.column_dimensions[get_column_letter(col)].width = 18

        # --- One sheet per (benchmark_type, case_name) when benchmark_type varies; else per case_name ---
        case_keys = list(
            OrderedDict.fromkeys(
                (d.get("benchmark_type", "Standard Performance"), d["case_name"]) for d in shown_data
            )
        )
        used_titles = set()
        for idx, (benchmark_type, case_name) in enumerate(case_keys):
            base = _sheet_title(f"{benchmark_type} | {case_name}", "Case") if len(benchmark_types) > 1 else _sheet_title(case_name, "Case")
            title = base
            n = 1
            while title in used_titles:
                suffix = f"_{n}"
                title = (base[: (EXCEL_SHEET_TITLE_MAX_LEN - len(suffix))] + suffix) if len(base) + len(suffix) > EXCEL_SHEET_TITLE_MAX_LEN else base + suffix
                n += 1
            used_titles.add(title)
            ws = wb.create_sheet(title, idx + 1)
            case_data = [
                d
                for d in shown_data
                if d["case_name"] == case_name and d.get("benchmark_type", "Standard Performance") == benchmark_type
            ]
            if not case_data:
                continue
            metrics_set = set()
            for d in case_data:
                metrics_set = metrics_set.union(d.get("metricsSet", set()))
            grouped = group_metrics_for_display(metrics_set)
            col = 1
            ws.cell(row=1, column=col, value="Instance")
            _style_header(ws.cell(row=1, column=col))
            _set_cell_comment(
                ws.cell(row=1, column=col),
                "Name of the database instance that produced this result (e.g. Clickhouse 1, Milvus 2).",
            )
            col += 1
            if len(benchmark_types) > 1:
                ws.cell(row=1, column=col, value="Benchmark Type")
                _style_header(ws.cell(row=1, column=col))
                _set_cell_comment(
                    ws.cell(row=1, column=col),
                    "Section from the Results page: Standard Performance, Int Filter, or Label Filter.",
                )
                col += 1
            for group_heading, metrics_in_group in grouped:
                for metric in metrics_in_group:
                    if any(_metric_value(d, metric) != "" for d in case_data):
                        unit = metric_unit_map.get(metric, "")
                        header = get_results_metric_display_name(metric)
                        # Size metrics are auto-scaled per cell (unit embedded in each value),
                        # so don't pin a single unit in the header.
                        if unit and _size_metric_base_bytes(metric) is None:
                            header = f"{header} ({unit})"
                        ws.cell(row=1, column=col, value=header)
                        _style_header(ws.cell(row=1, column=col))
                        desc = get_results_metric_tooltip(metric)
                        if desc:
                            _set_cell_comment(ws.cell(row=1, column=col), desc)
                        col += 1
            for r, d in enumerate(case_data, start=2):
                c = 1
                ws.cell(row=r, column=c, value=d.get("bar_display_name") or d.get("db_name", ""))
                _style_cell(ws.cell(row=r, column=c))
                c += 1
                if len(benchmark_types) > 1:
                    ws.cell(row=r, column=c, value=d.get("benchmark_type", ""))
                    _style_cell(ws.cell(row=r, column=c))
                    c += 1
                for _group_heading, metrics_in_group in grouped:
                    for metric in metrics_in_group:
                        if any(_metric_value(x, metric) != "" for x in case_data):
                            val = _metric_value(d, metric)
                            base_bytes = _size_metric_base_bytes(metric)
                            if base_bytes is not None and isinstance(val, (int, float)):
                                val = _human_readable_size(val * base_bytes)
                            ws.cell(row=r, column=c, value=val)
                            _style_cell(ws.cell(row=r, column=c))
                            c += 1
            for c in range(1, col + 1):
                ws.column_dimensions[get_column_letter(c)].width = 20

        # Per-component storage/RAM breakdown reported by the DB itself (currently Qdrant)
        _write_component_usage_sheet(wb, shown_data)

        # Theoretical best/expected/worst-case estimates for comparison against measured values
        _write_theoretical_sheet(wb, shown_data)

        # Glossary sheet: insert at index 1 so it appears right after Summary
        _write_glossary_sheet(wb)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()
